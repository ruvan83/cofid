from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE = Path('/mnt/data/work')
XLSX = Path('/mnt/data/McCance_Widdowsons_Composition_of_Foods_Integrated_Dataset_2021..xlsx')
OUT = BASE / 'app' / 'data'

CANONICAL = {
    'energy_kj': ['KJ', 'ENERGY (KJ) (KJ)', 'ENERGYKJ'],
    'energy_kcal': ['KCALS', 'ENERGY (KCAL) (KCAL)', 'ENERGYKCAL'],
    'protein_g': ['PROT', 'PROTEIN (G)', 'PROTEIN'],
    'fat_g': ['FAT', 'FAT (G)'],
    'carbohydrate_g': ['CHO', 'CARBOHYDRATE (G)', 'CARBOHYDRATE'],
    'sugars_g': ['TOTSUG', 'TOTAL SUGARS (G)', 'TOTAL SUGARS', 'SUGARS'],
    'fibre_g': ['AOACFIB', 'AOAC FIBRE (G)', 'AOAC FIBRE', 'FIBRE'],
    'sodium_mg': ['SODIUM', 'SODIUM (MG)'],
    'salt_g': ['SALT'],
    'saturates_g': ['SATFOD', 'SATURATED FATTY ACIDS PER 100G FOOD', 'SATURATES'],
}


def normalize_key(value: Any) -> str:
    return re.sub(r'[^A-Z0-9]+', '', str(value or '').upper())


def clean_value(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if v in {'', 'N'}:
            return None
        if v == 'Tr':
            return 0.0
        value = v
    try:
        value = float(value)
    except Exception:
        return None
    if math.isnan(value):
        return None
    return value


def build():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    prox = wb['1.3 Proximates']

    header1 = [cell for cell in next(prox.iter_rows(min_row=1, max_row=1, values_only=True))]
    header2 = [cell for cell in next(prox.iter_rows(min_row=2, max_row=2, values_only=True))]
    header3 = [cell for cell in next(prox.iter_rows(min_row=3, max_row=3, values_only=True))]

    columns = []
    for i in range(len(header1)):
        columns.append({
            'display': str(header1[i] or header3[i] or header2[i] or f'col_{i}').strip(),
            'code': str(header2[i] or '').strip(),
            'fallback': str(header3[i] or '').strip(),
            'index': i,
        })

    def find_index(candidates):
        normalized = {}
        for col in columns:
            normalized[normalize_key(col['display'])] = col['index']
            if col['code']:
                normalized[normalize_key(col['code'])] = col['index']
            if col['fallback']:
                normalized[normalize_key(col['fallback'])] = col['index']
        for cand in candidates:
            key = normalize_key(cand)
            if key in normalized:
                return normalized[key]
        return None

    food_code_idx = find_index(['Food Code'])
    name_idx = find_index(['Food Name'])
    desc_idx = find_index(['Description'])
    group_idx = find_index(['Group'])

    canonical = {}
    for key, candidates in CANONICAL.items():
        idx = find_index(candidates)
        if idx is not None:
            canonical[key] = idx

    nutrient_defs = []
    for key, idx in canonical.items():
        unit = 'g'
        if key.endswith('_kj'):
            unit = 'kJ'
        elif key.endswith('_kcal'):
            unit = 'kcal'
        elif key.endswith('_mg'):
            unit = 'mg'
        nutrient_defs.append({
            'key': key,
            'label': columns[idx]['display'],
            'unit': unit,
        })

    foods = []
    for row in prox.iter_rows(min_row=4, values_only=True):
        if food_code_idx is None or name_idx is None:
            raise RuntimeError('Required columns not found.')
        food_code = row[food_code_idx] if food_code_idx < len(row) else None
        name = row[name_idx] if name_idx < len(row) else None
        if food_code is None or name in (None, ''):
            continue
        nutrients = {}
        for key, idx in canonical.items():
            if idx < len(row):
                val = clean_value(row[idx])
                if val is not None:
                    nutrients[key] = val
        foods.append({
            'food_code': str(food_code).strip(),
            'name': str(name).strip(),
            'description': '' if desc_idx is None or desc_idx >= len(row) or row[desc_idx] is None else str(row[desc_idx]).strip(),
            'group': '' if group_idx is None or group_idx >= len(row) or row[group_idx] is None else str(row[group_idx]).strip(),
            'nutrients': nutrients,
        })

    meta = {
        'source': 'McCance and Widdowson CoFID',
        'workbook': XLSX.name,
        'food_count': len(foods),
        'nutrients': nutrient_defs,
        'warning': 'Reference values are typical composition data. Match foods carefully, especially for processed foods and alcoholic drinks. Alcoholic drinks in CoFID may be expressed per 100 ml in the source workbook.',
    }
    return meta, foods


def main():
    meta, foods = build()
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / 'meta.json').write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
    (OUT / 'foods.json').write_text(json.dumps(foods, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
    (OUT / 'meta.js').write_text('window.COFID_EMBEDDED_META = ' + json.dumps(meta, ensure_ascii=False, separators=(',', ':')) + ';\n', encoding='utf-8')
    (OUT / 'foods.js').write_text('window.COFID_EMBEDDED_FOODS = ' + json.dumps(foods, ensure_ascii=False, separators=(',', ':')) + ';\n', encoding='utf-8')
    print(f'foods={len(foods)}')

if __name__ == '__main__':
    main()
